#!/usr/bin/env python3
"""AI CCTV Backend - Entry Point"""
from pathlib import Path
from flask import Flask, send_file, send_from_directory, jsonify, request, Response, g
from flask_cors import CORS
from core import (
    Config, Database, MLService, VideoCamera,
    RateLimiter, Validator, ValidationError, AuthenticationError,
    AuthorizationError, NotFoundError, APIError,
    logger, DEVICE,
    save_detection_to_db, update_inventory_from_detection,
    check_low_stock_alerts, create_alert,
    create_token, decode_token, token_required, admin_required, rate_limit,
    ml_service, db, rate_limiter, DetectionResult,
    normalize_camera_source, camera_source_kind,
    generate_password_hash, check_password_hash,
)
import threading, time, uuid, csv, io, os, cv2, numpy as np, subprocess, json
from datetime import datetime, timedelta
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any, Tuple
from enum import Enum
from pathlib import Path as _Path
import jwt

FRONTEND_DIST = Config.FRONTEND_DIST

camera = None  # global live camera instance

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = Config.MAX_UPLOAD_SIZE
CORS(app, resources={r"/*": {"origins": "*"}}, max_age=3600)  # cache preflight for 1 h

@app.before_request
def handle_preflight():
    if request.method == "OPTIONS":
        response = app.make_default_options_response()
        response.headers["Access-Control-Allow-Origin"] = "*"
        response.headers["Access-Control-Allow-Headers"] = "*"
        response.headers["Access-Control-Allow-Methods"] = "*"
        return response

@app.errorhandler(APIError)
def handle_api_error(error):
    return jsonify(error.to_dict()), error.status_code

@app.teardown_appcontext
def teardown_db(exception):
    db.close()

# ============================================================================
# ROUTES: CORE
# ============================================================================

@app.route('/api/info')
def index():
    """Backend metadata — moved to /api/info so '/' serves the React frontend."""
    return jsonify({'name': 'AI CCTV Edge', 'version': '4.1-edge', 'status': 'running', 'mode': 'edge', 'gpu': DEVICE == 'cuda'})


@app.route('/health')
@app.route('/api/health')
def health():
    return jsonify({'status': 'ok', 'mode': 'edge', **ml_service.health_check()})

# ============================================================================
# ROUTES: AUTH
# ============================================================================

@app.route('/api/auth/login', methods=['POST'])
@rate_limit
def login():
    data = request.json or {}
    Validator(data).require('email').email('email').require('password').validate()
    user = db.fetchone("SELECT * FROM users WHERE email = ? AND is_active = 1", (data['email'],))
    if not user or not check_password_hash(user['password_hash'], data['password']):
        raise AuthenticationError("Invalid email or password")
    token = create_token(user['id'], user['role'], user['name'])
    return jsonify({'access_token': token, 'user': {'id': user['id'], 'email': user['email'], 'name': user['name'], 'role': user['role']}})

# ============================================================================
# ROUTES: MODELS
# ============================================================================

@app.route('/api/models', methods=['GET'])
@token_required
def list_models():
    return jsonify({'available': ml_service.loaded_models, 'active': ml_service.active_model_name})

@app.route('/api/models/switch', methods=['POST'])
@token_required
def switch_model():
    name = (request.json or {}).get('model')
    if ml_service.switch_model(name): return jsonify({'active': name})
    raise NotFoundError("Model")

# ============================================================================
# ROUTES: DETECTION
# ============================================================================

@app.route('/api/detect_frame', methods=['POST'])
@token_required
def detect_frame():
    if 'image' not in request.files: raise ValidationError("No image provided")
    
    file_bytes = np.frombuffer(request.files['image'].read(), np.uint8)
    image = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
    if image is None: raise ValidationError("Invalid image")
    
    save = request.form.get('save', 'false').lower() == 'true'
    direction = request.form.get('direction')
    confidence = float(request.form.get('confidence', Config.ML_DEFAULT_CONFIDENCE))
    
    _, detections = ml_service.detect(image, confidence=confidence, draw=False)
    
    if save:
        for det in detections[:10]:
            save_detection_to_db(det.label, det.confidence, direction, det.inference_ms)
            if direction: update_inventory_from_detection(det.label, direction)
    
    return jsonify({'detections': [d.to_dict() for d in detections], 'count': len(detections), 'saved': save})

@app.route('/api/detections/live')
@token_required
def live_detections():
    return jsonify(ml_service.latest_detections)

@app.route('/api/detections')
@token_required
def get_detections():
    limit = min(int(request.args.get('limit', 50)), 200)
    dets = db.fetchall("SELECT * FROM detections ORDER BY detected_at DESC LIMIT ?", (limit,))
    return jsonify([dict(d) for d in dets])

@app.route('/api/detections/record', methods=['POST'])
@token_required
def record_detection():
    data = request.json or {}
    Validator(data).require('type').require('direction').in_list('direction', ['IN', 'OUT']).validate()
    
    detection_id = str(uuid.uuid4())
    db.insert('detections', {
        'id': detection_id,
        'type': data['type'],
        'confidence': data.get('confidence', 1.0),
        'direction': data['direction'],
        'inference_ms': data.get('inference_ms', 0)
    })
    update_inventory_from_detection(data['type'], data['direction'], data.get('count', 1))
    
    return jsonify({'id': detection_id, 'message': f"Detection recorded ({data['direction']})"}), 201

# ============================================================================
# ROUTES: STATS (REAL DATA)
# ============================================================================

@app.route('/api/stats')
@token_required
def get_stats():
    inv = db.fetchone("SELECT COALESCE(SUM(count_in),0) as total_in, COALESCE(SUM(count_out),0) as total_out, COALESCE(SUM(current_stock),0) as total_stock FROM inventory")
    today = datetime.now().strftime('%Y-%m-%d')
    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    
    today_stats = db.fetchone("""
        SELECT COALESCE(SUM(CASE WHEN direction='IN' THEN 1 ELSE 0 END),0) as today_in,
               COALESCE(SUM(CASE WHEN direction='OUT' THEN 1 ELSE 0 END),0) as today_out
        FROM detections WHERE DATE(detected_at) = ?
    """, (today,))
    
    yesterday_stats = db.fetchone("""
        SELECT COALESCE(SUM(CASE WHEN direction='IN' THEN 1 ELSE 0 END),0) as yesterday_in,
               COALESCE(SUM(CASE WHEN direction='OUT' THEN 1 ELSE 0 END),0) as yesterday_out
        FROM detections WHERE DATE(detected_at) = ?
    """, (yesterday,))
    
    y_in = yesterday_stats['yesterday_in'] if yesterday_stats else 0
    y_out = yesterday_stats['yesterday_out'] if yesterday_stats else 0
    t_in = today_stats['today_in'] if today_stats else 0
    t_out = today_stats['today_out'] if today_stats else 0
    
    in_pct = ((t_in - y_in) / max(1, y_in)) * 100
    out_pct = ((t_out - y_out) / max(1, y_out)) * 100
    
    in_trend = f"{in_pct:+.1f}%"
    out_trend = f"{out_pct:+.1f}%"
    
    sugar = db.fetchone("SELECT COALESCE(current_stock,0) as count FROM inventory WHERE LOWER(product_name) LIKE '%sugar%'")
    
    trucks = db.fetchone("""
        SELECT COALESCE(SUM(CASE WHEN direction='IN' THEN 1 ELSE 0 END),0) as trucks_in,
               COALESCE(SUM(CASE WHEN direction='OUT' THEN 1 ELSE 0 END),0) as trucks_out
        FROM trucks WHERE DATE(detected_at) = ?
    """, (today,))
    
    return jsonify({
        'total_in': inv['total_in'], 'total_out': inv['total_out'], 'total_stock': inv['total_stock'],
        'today_in': today_stats['today_in'] if today_stats else 0,
        'today_out': today_stats['today_out'] if today_stats else 0,
        'in_trend': in_trend, 'out_trend': out_trend,
        'in_up': in_pct >= 0, 'out_up': out_pct >= 0,
        'sugar_bag_count': sugar['count'] if sugar else 0,
        'trucks_in_today': trucks['trucks_in'] if trucks else 0,
        'trucks_out_today': trucks['trucks_out'] if trucks else 0,
        'gpu_active': DEVICE == 'cuda',
        'active_model': ml_service.active_model_name,
        'timestamp': datetime.now().isoformat()
    })


# ============================================================================
# ROUTES: ANALYTICS (REAL DATA)
# ============================================================================

@app.route('/api/analytics')
@token_required
def get_analytics():
    today = datetime.now().strftime('%Y-%m-%d')
    days = int(request.args.get('days', 7))
    
    hourly_data = db.fetchall("""
        SELECT CAST(strftime('%H', detected_at) AS INTEGER) as hour,
               SUM(CASE WHEN direction='IN' THEN 1 ELSE 0 END) as in_count,
               SUM(CASE WHEN direction='OUT' THEN 1 ELSE 0 END) as out_count,
               COUNT(*) as total
        FROM detections WHERE DATE(detected_at) = ?
        GROUP BY hour ORDER BY hour
    """, (today,))
    
    hourly_dict = {row['hour']: row for row in hourly_data}
    hourly = [{'hour': h, 'in_count': hourly_dict[h]['in_count'] if h in hourly_dict else 0, 'out_count': hourly_dict[h]['out_count'] if h in hourly_dict else 0, 'total': hourly_dict[h]['total'] if h in hourly_dict else 0} for h in range(24)]
    
    daily_data = db.fetchall("""
        SELECT DATE(detected_at) as date,
               SUM(CASE WHEN direction='IN' THEN 1 ELSE 0 END) as in_count,
               SUM(CASE WHEN direction='OUT' THEN 1 ELSE 0 END) as out_count,
               COUNT(*) as total
        FROM detections WHERE detected_at >= DATE('now', ?)
        GROUP BY DATE(detected_at) ORDER BY date
    """, (f'-{days} days',))
    
    type_dist = db.fetchall("SELECT type, COUNT(*) as count FROM detections WHERE detected_at >= DATE('now', '-7 days') GROUP BY type ORDER BY count DESC LIMIT 10")
    
    peak = db.fetchone("SELECT CAST(strftime('%H', detected_at) AS INTEGER) as hour, COUNT(*) as count FROM detections WHERE DATE(detected_at) = ? GROUP BY hour ORDER BY count DESC LIMIT 1", (today,))
    
    return jsonify({
        'hourly': hourly,
        'daily': [dict(d) for d in daily_data],
        'type_distribution': [dict(t) for t in type_dist],
        'summary': {'peak_hour': peak['hour'] if peak else None, 'peak_count': peak['count'] if peak else 0}
    })

# ============================================================================
# ROUTES: INVENTORY
# ============================================================================

@app.route('/api/inventory')
@token_required
def get_inventory():
    items = db.fetchall("SELECT * FROM inventory ORDER BY product_name")
    return jsonify([dict(i) for i in items])

@app.route('/api/inventory', methods=['POST'])
@admin_required
def add_inventory():
    data = request.json or {}
    Validator(data).require('product_name').validate()
    if db.fetchone("SELECT id FROM inventory WHERE product_name = ?", (data['product_name'],)):
        raise ValidationError("Product already exists")
    item_id = str(uuid.uuid4())
    db.insert('inventory', {'id': item_id, 'product_name': data['product_name'], 'min_threshold': data.get('min_threshold', 10)})
    return jsonify({'id': item_id}), 201

@app.route('/api/inventory/<item_id>/update', methods=['POST'])
@token_required
def update_inventory_count(item_id):
    data = request.json or {}
    Validator(data).require('direction').in_list('direction', ['IN', 'OUT']).validate()
    count = int(data.get('count', 1))
    direction = data['direction']
    
    item = db.fetchone("SELECT * FROM inventory WHERE id = ?", (item_id,))
    if not item: raise NotFoundError("Inventory item")
    
    with db.get_cursor() as cursor:
        if direction == 'IN':
            cursor.execute("UPDATE inventory SET count_in = count_in + ?, current_stock = current_stock + ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (count, count, item_id))
        else:
            cursor.execute("UPDATE inventory SET count_out = count_out + ?, current_stock = MAX(0, current_stock - ?), updated_at = CURRENT_TIMESTAMP WHERE id = ?", (count, count, item_id))
    
    return jsonify({'message': f'Updated {direction}', 'count': count})

# ============================================================================
# ROUTES: TRUCKS
# ============================================================================

@app.route('/api/trucks')
@token_required
def get_trucks():
    trucks = db.fetchall("SELECT * FROM trucks ORDER BY detected_at DESC LIMIT 50")
    return jsonify([dict(t) for t in trucks])

@app.route('/api/trucks', methods=['POST'])
@token_required
def add_truck():
    data = request.json or {}
    plate = data.get('truck_number') or data.get('plate_number')
    if not plate: raise ValidationError("Plate number required")
    tid = str(uuid.uuid4())
    db.insert('trucks', {'id': tid, 'plate_number': plate.upper(), 'direction': data.get('direction', 'IN'), 'driver_name': data.get('driver_name'), 'company': data.get('company'), 'purpose': data.get('purpose')})
    return jsonify({'id': tid}), 201

@app.route('/api/trucks/<truck_id>/exit', methods=['POST'])
@token_required
def truck_exit(truck_id):
    truck = db.fetchone("SELECT * FROM trucks WHERE id = ?", (truck_id,))
    if not truck: raise NotFoundError("Truck")
    db.update('trucks', {'direction': 'OUT', 'exit_at': datetime.now().isoformat()}, 'id = ?', (truck_id,))
    return jsonify({'message': 'Truck exit recorded'})

@app.route('/api/trucks/reset', methods=['DELETE'])
@token_required
def reset_trucks():
    count = db.delete('trucks')
    return jsonify({'deleted': count})

# ============================================================================
# ROUTES: FACES
# ============================================================================

@app.route('/api/faces')
@token_required
def get_faces():
    faces = db.fetchall("SELECT id, name, created_at FROM faces ORDER BY created_at DESC")
    return jsonify([dict(f) for f in faces])

@app.route('/api/faces', methods=['POST'])
@token_required
def add_face():
    data = request.json or {}
    Validator(data).require('name').validate()
    with db.get_cursor() as cursor:
        cursor.execute("INSERT INTO faces (name) VALUES (?)", (data['name'],))
        face_id = cursor.lastrowid
    return jsonify({'id': face_id, 'message': 'Face registered'}), 201

@app.route('/api/faces/<int:face_id>', methods=['DELETE'])
@token_required
def delete_face(face_id):
    if db.delete('faces', 'id = ?', (face_id,)) == 0: raise NotFoundError("Face")
    return jsonify({'message': 'Face deleted'})

# ============================================================================
# ROUTES: SCANS
# ============================================================================

@app.route('/api/scans')
@token_required
def get_scans():
    scans = db.fetchall("SELECT * FROM scans ORDER BY scanned_at DESC LIMIT 100")
    return jsonify([dict(s) for s in scans])

@app.route('/api/scans', methods=['POST'])
@token_required
def add_scan():
    data = request.json or {}
    # Require either barcode or batch_number
    if not data.get('barcode') and not data.get('batch_number'):
        raise ValidationError("Either barcode or batch_number is required")
    
    exp_date = data.get('exp_date', '')
    is_expired = False
    days_until = None
    if exp_date:
        try:
            # Try multiple date formats
            for fmt in ['%Y-%m-%d', '%d/%m/%y', '%d/%m/%Y', '%d-%m-%Y']:
                try:
                    exp = datetime.strptime(exp_date, fmt)
                    days_until = (exp - datetime.now()).days
                    is_expired = days_until < 0
                    break
                except ValueError:
                    continue
        except: pass
    
    scan_id = str(uuid.uuid4())
    db.insert('scans', {
        'id': scan_id, 
        'barcode': data.get('barcode', ''), 
        'product_name': data.get('product_name', ''),
        'batch_number': data.get('batch_number', ''), 
        'mfg_date': data.get('mfg_date', ''),
        'exp_date': exp_date, 
        'rack_no': data.get('rack_no', ''), 
        'shelf_no': data.get('shelf_no', ''),
        'direction': data.get('direction', 'IN')
    })
    return jsonify({'id': scan_id, 'is_expired': is_expired, 'days_until_expiry': days_until}), 201

@app.route('/api/scans/<scan_id>', methods=['DELETE'])
@token_required
def delete_scan(scan_id):
    if db.delete('scans', 'id = ?', (scan_id,)) == 0: raise NotFoundError("Scan")
    return jsonify({'message': 'Scan deleted'})

@app.route('/api/barcode/decode', methods=['POST'])
@token_required
def decode_barcode():
    if not BARCODE_AVAILABLE: raise APIError("Barcode decoding not available", 503)
    if 'image' not in request.files: raise ValidationError("No image provided")
    
    file_bytes = np.frombuffer(request.files['image'].read(), np.uint8)
    image = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
    if image is None: raise ValidationError("Invalid image")
    
    barcodes = pyzbar.decode(image)
    results = []
    for bc in barcodes:
        x, y, w, h = bc.rect
        results.append({'data': bc.data.decode('utf-8'), 'type': bc.type, 'bbox': [x, y, x+w, y+h]})
    return jsonify({'barcodes': results, 'count': len(results)})

# ============================================================================
# ROUTES: OCR
# ============================================================================

# Try to import pytesseract for OCR
try:
    import pytesseract
    from PIL import Image
    import io as io_module
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

# Known beverage flavors for OCR detection
KNOWN_FLAVORS = [
    'PEPSI', 'COLA', 'SPRITE', 'FANTA', '7UP', '7 UP', 'MIRINDA',
    'MOUNTAIN DEW', 'DEW', 'SLICE', 'MAAZA', 'FROOTI', 'APPY',
    'LIMCA', 'THUMS UP', 'THUMBS UP', 'MANGO', 'ORANGE', 'LEMON',
    'STING', 'GATORADE', 'TROPICANA', 'AQUAFINA', 'KINLEY',
    'NIMBOOZ', 'TROPICANA', 'REAL', 'PAPER BOAT', 'RAW', 'TROPICO'
]

def find_flavour(text: str) -> str:
    """Find product flavour from OCR text"""
    upper = text.upper()
    
    for f in KNOWN_FLAVORS:
        if f in upper:
            # Title case the flavor
            return ' '.join(w.capitalize() for w in f.split())
    
    # Try to find "FLAVOR" or "FLAVOUR" followed by product name
    flavor_match = re.search(r'(?:FLAVOR|FLAVOUR)\s+(\w+)', upper)
    if flavor_match:
        return flavor_match.group(1).capitalize()
    
    return ''

def find_all_dates(text: str) -> list:
    """Find all dates in various formats"""
    dates = []
    # DD/MM/YY or DD/MM/YYYY (also handles . and - separators)
    regex = r'(\d{1,2})[\/.\-](\d{1,2})[\/.\-](\d{2,4})'
    for match in re.finditer(regex, text):
        raw = match.group(0)
        # Normalize date
        d, mo, y = match.groups()
        if len(y) == 2:
            y = '20' + y
        normalized = f"{d.zfill(2)}/{mo.zfill(2)}/{y}"
        dates.append({
            'raw': raw,
            'normalized': normalized,
            'index': match.start()
        })
    return dates

def find_batch(text: str) -> str:
    """Find batch number - optimized for XX-XXXX-XXXX format"""
    upper = text.upper()
    
    patterns = [
        r'(\d{2}-\d{4}-\d{4})',  # 25-8902-0014 (exact format)
        r'BATCH\s*NO\.?\s*[:\s]*([A-Z0-9\-]{6,})',
        r'B\.?\s*NO\.?\s*[:\s]*([A-Z0-9\-]{6,})',
        r'LOT\s*(?:NO\.?)?\s*[:\s]*([A-Z0-9\-]{5,})',
    ]
    
    for p in patterns:
        match = re.search(p, upper)
        if match:
            batch = match.group(1).strip()
            # Make sure it's not a date
            if not re.match(r'^\d{1,2}[\/.\-]\d{1,2}[\/.\-]\d{2,4}$', batch):
                return batch
    
    return ''

def parse_dates_from_context(text: str, dates: list) -> dict:
    """Parse dates based on context (MFG/EXP keywords position)"""
    upper = text.upper()
    result = {'mfg': '', 'expiry': ''}
    
    if not dates:
        return result
    
    # Find keyword positions
    mfg_keywords = ['MANUFACTURE DATE', 'MFG DATE', 'MFG DT', 'MFD', 'PACKED', 'PKD']
    exp_keywords = ['EXPIRY DATE', 'EXP DATE', 'EXP DT', 'BEST BEFORE', 'USE BY', 'BB']
    
    mfg_pos = -1
    for kw in mfg_keywords:
        pos = upper.find(kw)
        if pos != -1 and (mfg_pos == -1 or pos < mfg_pos):
            mfg_pos = pos
    
    exp_pos = -1
    for kw in exp_keywords:
        pos = upper.find(kw)
        if pos != -1 and (exp_pos == -1 or pos < exp_pos):
            exp_pos = pos
    
    # If MFG keyword found and we have at least 2 dates
    if mfg_pos != -1 and len(dates) >= 2:
        dates_after = [d for d in dates if d['index'] > mfg_pos]
        if len(dates_after) >= 2:
            result['mfg'] = dates_after[0]['normalized']
            result['expiry'] = dates_after[1]['normalized']
            return result
    
    # Fallback: use date order (first = MFG, second = EXP)
    if len(dates) >= 2:
        result['mfg'] = dates[0]['normalized']
        result['expiry'] = dates[1]['normalized']
    elif len(dates) == 1:
        if exp_pos != -1 and (mfg_pos == -1 or exp_pos < mfg_pos):
            result['expiry'] = dates[0]['normalized']
        else:
            result['mfg'] = dates[0]['normalized']
    
    return result

def parse_label_text(text: str) -> dict:
    """Parse OCR text to extract batch, expiry, MFG date, and flavour - optimized for PepsiCo labels"""
    result = {
        'batch_no': '',
        'expiry_date': '',
        'mfg_date': '',
        'flavour': '',
        'raw_text': text[:300] if text else ''
    }
    
    if not text:
        return result
    
    # Find all dates
    dates = find_all_dates(text)
    
    # Extract batch number
    result['batch_no'] = find_batch(text)
    
    # Extract dates based on context
    date_info = parse_dates_from_context(text, dates)
    result['mfg_date'] = date_info['mfg']
    result['expiry_date'] = date_info['expiry']
    
    # Extract flavour
    result['flavour'] = find_flavour(text)
    
    # Validate: swap dates if expiry < mfg
    if result['mfg_date'] and result['expiry_date']:
        try:
            mfg_parts = result['mfg_date'].split('/')
            exp_parts = result['expiry_date'].split('/')
            mfg_date = datetime(int(mfg_parts[2]), int(mfg_parts[1]), int(mfg_parts[0]))
            exp_date = datetime(int(exp_parts[2]), int(exp_parts[1]), int(exp_parts[0]))
            if exp_date < mfg_date:
                result['mfg_date'], result['expiry_date'] = result['expiry_date'], result['mfg_date']
        except:
            pass
    
    return result

@app.route('/api/ocr/scan', methods=['POST'])
@token_required
def ocr_scan():
    """OCR scan an image to extract label data"""
    if not OCR_AVAILABLE:
        raise APIError("OCR not available. Install pytesseract and pillow.", 503)
    
    if 'image' not in request.files:
        raise ValidationError("No image provided")
    
    try:
        file = request.files['image']
        img = Image.open(io_module.BytesIO(file.read()))
        
        # Run OCR
        text = pytesseract.image_to_string(img)
        result = parse_label_text(text)
        
        return jsonify({
            'success': True,
            'batch_no': result['batch_no'],
            'mfg_date': result['mfg_date'],
            'expiry_date': result['expiry_date'],
            'flavour': result['flavour'],
            'raw_text': result['raw_text']
        })
    except Exception as e:
        log(f"OCR error: {e}", "ERROR")
        return jsonify({
            'success': False,
            'error': str(e),
            'batch_no': '',
            'mfg_date': '',
            'expiry_date': '',
            'flavour': ''
        })

# ============================================================================
# ROUTES: ALERTS
# ============================================================================

@app.route('/api/alerts')
@token_required
def get_alerts():
    unread = request.args.get('unread', 'false').lower() == 'true'
    limit = min(int(request.args.get('limit', 50)), 100)
    q = "SELECT * FROM alerts WHERE is_read = 0 ORDER BY created_at DESC LIMIT ?" if unread else "SELECT * FROM alerts ORDER BY created_at DESC LIMIT ?"
    return jsonify([dict(a) for a in db.fetchall(q, (limit,))])

@app.route('/api/alerts/<alert_id>/read', methods=['POST'])
@token_required
def mark_alert_read(alert_id):
    if db.update('alerts', {'is_read': 1}, 'id = ?', (alert_id,)) == 0: raise NotFoundError("Alert")
    return jsonify({'message': 'Alert marked as read'})

@app.route('/api/alerts/read-all', methods=['POST'])
@token_required
def mark_all_read():
    db.update('alerts', {'is_read': 1}, '1=1', ())
    return jsonify({'message': 'All alerts marked as read'})

# ============================================================================
# ROUTES: CAMERA
# ============================================================================

@app.route('/api/camera/start', methods=['POST'])
@token_required
def start_camera():
    global camera
    if camera is not None:
        return jsonify({'status': 'already running', 'source': camera.source})
    data = request.get_json(silent=True) or {}
    source = normalize_camera_source(data.get('source', 0))
    try:
        camera = VideoCamera(source, camera_id='live')
        return jsonify({'status': 'started', 'source': str(source), 'source_kind': camera.source_kind})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/camera/stop', methods=['POST'])
@token_required
def stop_camera():
    global camera
    if camera:
        camera.stop()
        camera = None
    return jsonify({'status': 'stopped'})

@app.route('/api/camera/status')
@token_required
def camera_status():
    """Returns live session crossing counts, line position, confidence and frame_skip."""
    if camera is None:
        return jsonify({'active': False, 'session_in': 0, 'session_out': 0,
                        'line_position': 0.45, 'confidence': 0.5, 'frame_skip': 3})
    return jsonify({
        'active': True,
        'session_in': camera.session_in,
        'session_out': camera.session_out,
        'line_position': camera.line_position,
        'confidence': camera.confidence,
        'frame_skip': camera.frame_skip,
        'source': str(camera.source),
        'source_kind': camera.source_kind,
    })


@app.route('/api/camera/settings', methods=['PATCH'])
@token_required
def set_camera_settings():
    """Update confidence and/or frame_skip on the live camera without restarting."""
    if camera is None:
        return jsonify({'error': 'Camera is not running'}), 400
    data = request.get_json(silent=True) or {}
    if 'confidence' in data:
        val = float(data['confidence'])
        if not (0.05 <= val <= 1.0):
            return jsonify({'error': 'confidence must be between 0.05 and 1.0'}), 400
        camera.confidence = val
    if 'frame_skip' in data:
        val = int(data['frame_skip'])
        if not (1 <= val <= 30):
            return jsonify({'error': 'frame_skip must be between 1 and 30'}), 400
        camera.frame_skip = val
    return jsonify({'confidence': camera.confidence, 'frame_skip': camera.frame_skip})

@app.route('/api/camera/line', methods=['PATCH'])
@token_required
def set_camera_line():
    """Update tracking line position (0.0–1.0) without restarting the camera."""
    data = request.get_json(silent=True) or {}
    pos = data.get('position')
    if pos is None or not (0.0 <= float(pos) <= 1.0):
        return jsonify({'error': 'position must be a float between 0.0 and 1.0'}), 400
    if camera is None:
        return jsonify({'error': 'Camera is not running'}), 400
    camera.line_position = float(pos)
    return jsonify({'line_position': camera.line_position})


@app.route('/api/video_feed')
def video_feed():
    """MJPEG stream with optional ?quality=low for remote/slow-connection viewers.
    normal: 1280px, Q70, 15fps  ≈ 3–8 Mbps
    low:     480px,  Q40,  6fps  ≈ 0.3–1 Mbps (works on a 1 Mbps upload)
    """
    low = request.args.get('quality') == 'low'
    fps   = 6   if low else 15
    width = 480 if low else 1280
    qual  = 40  if low else 70

    def gen():
        interval = 1.0 / fps
        while camera:
            t0 = time.time()
            frame = camera.get_frame()
            if frame is not None:
                h, w = frame.shape[:2]
                if w > width:
                    frame = cv2.resize(frame, (width, int(h * width / w)), interpolation=cv2.INTER_LINEAR)
                ok, buf = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, qual])
                if ok:
                    yield b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buf.tobytes() + b'\r\n'
            elapsed = time.time() - t0
            rem = interval - elapsed
            if rem > 0:
                time.sleep(rem)
    return Response(gen(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/api/video_feed/low')
def video_feed_low():
    """Convenience alias for ?quality=low — link this in the UI for remote viewers."""
    from flask import redirect
    return redirect('/api/video_feed?quality=low')



# ============================================================================
# ROUTES: BAYS  (multi-camera loading dock management)
# ============================================================================

@app.route('/api/bays', methods=['GET'])
@token_required
def list_bays():
    bays = db.fetchall("SELECT * FROM bays ORDER BY name")
    return jsonify({'bays': bays})

@app.route('/api/bays', methods=['POST'])
@token_required
def create_bay():
    d = request.get_json(silent=True) or {}
    if not d.get('name'):
        raise ValidationError("name is required")
    bay_id = str(uuid.uuid4())
    db.insert('bays', {
        'id': bay_id,
        'name': d['name'],
        'camera_source': str(normalize_camera_source(d.get('camera_source', 0))),
        'location': d.get('location', ''),
        'is_active': 1,
    })
    return jsonify({'id': bay_id, 'message': 'Bay created'}), 201

@app.route('/api/bays/<bay_id>', methods=['PATCH'])
@token_required
def update_bay(bay_id):
    d = request.get_json(silent=True) or {}
    fields = {k: v for k, v in d.items() if k in ('name', 'camera_source', 'location', 'is_active')}
    if 'camera_source' in fields:
        fields['camera_source'] = str(normalize_camera_source(fields['camera_source']))
    if fields:
        if db.update('bays', fields, 'id = ?', (bay_id,)) == 0:
            raise NotFoundError("Bay")
    return jsonify({'message': 'Updated'})

@app.route('/api/bays/<bay_id>', methods=['DELETE'])
@token_required
def delete_bay(bay_id):
    if db.delete('bays', 'id = ?', (bay_id,)) == 0:
        raise NotFoundError("Bay")
    return jsonify({'message': 'Deleted'})

@app.route('/api/bays/<bay_id>/feed')
def bay_video_feed(bay_id):
    """MJPEG stream for a specific bay camera."""
    bay = db.fetchone("SELECT * FROM bays WHERE id=?", (bay_id,))
    if not bay:
        return jsonify({'error': 'Bay not found'}), 404
    src = normalize_camera_source(bay['camera_source'])
    cam = camera_manager.get_or_create(bay_id, src)

    def gen():
        interval = 1.0 / 15
        while True:
            t0 = time.time()
            buf = cam.get_encoded_frame()
            if buf:
                yield b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + buf + b'\r\n'
            elapsed = time.time() - t0
            rem = interval - elapsed
            if rem > 0:
                time.sleep(rem)
    return Response(gen(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/api/bays/<bay_id>/start', methods=['POST'])
@token_required
def start_bay_camera(bay_id):
    bay = db.fetchone("SELECT * FROM bays WHERE id=?", (bay_id,))
    if not bay:
        raise ValidationError("Bay not found")
    src = normalize_camera_source(bay['camera_source'])
    cam = camera_manager.get_or_create(bay_id, src)
    return jsonify({'status': 'started', 'bay_id': bay_id, 'source': str(src), 'source_kind': cam.source_kind})

@app.route('/api/bays/<bay_id>/stop', methods=['POST'])
@token_required
def stop_bay_camera(bay_id):
    camera_manager.stop(bay_id)
    return jsonify({'status': 'stopped', 'bay_id': bay_id})

@app.route('/api/bays/<bay_id>/status')
@token_required
def bay_camera_status(bay_id):
    cam = camera_manager.cameras.get(bay_id)
    if cam is None:
        return jsonify({'active': False, 'session_in': 0, 'session_out': 0, 'line_position': 0.45})
    return jsonify({'active': True, 'session_in': cam.session_in, 'session_out': cam.session_out,
                    'line_position': cam.line_position, 'source': str(cam.source),
                    'source_kind': cam.source_kind})


# ============================================================================
# ROUTES: SHIFTS
# ============================================================================

@app.route('/api/shifts', methods=['GET'])
@token_required
def list_shifts():
    shifts = db.fetchall(
        "SELECT s.*, "
        "(SELECT COUNT(*) FROM detections d WHERE d.shift_id=s.id) as detection_count "
        "FROM shifts s ORDER BY s.started_at DESC LIMIT 50"
    )
    return jsonify({'shifts': shifts})

@app.route('/api/shifts/active', methods=['GET'])
@token_required
def get_active_shift():
    shift = db.fetchone("SELECT * FROM shifts WHERE ended_at IS NULL ORDER BY started_at DESC LIMIT 1")
    return jsonify({'shift': shift})

@app.route('/api/shifts/start', methods=['POST'])
@token_required
def start_shift():
    d = request.get_json(silent=True) or {}
    # Auto-close any open shift first
    db.execute("UPDATE shifts SET ended_at=CURRENT_TIMESTAMP WHERE ended_at IS NULL")
    shift_id = str(uuid.uuid4())
    db.insert('shifts', {
        'id': shift_id,
        'name': d.get('name', f"Shift {shift_id[:6]}"),
        'operator_name': d.get('operator_name', ''),
        'notes': d.get('notes', ''),
    })
    logger.info(f"🕐 Shift started: {shift_id}")
    return jsonify({'id': shift_id, 'message': 'Shift started'}), 201

@app.route('/api/shifts/end', methods=['POST'])
@token_required
def end_shift():
    d = request.get_json(silent=True) or {}
    shift = db.fetchone("SELECT * FROM shifts WHERE ended_at IS NULL ORDER BY started_at DESC LIMIT 1")
    if not shift:
        return jsonify({'error': 'No active shift'}), 400
    db.execute("UPDATE shifts SET ended_at=CURRENT_TIMESTAMP, notes=? WHERE id=?",
               (d.get('notes', shift.get('notes', '')), shift['id']))
    # Close any open jobs for this shift
    db.execute("UPDATE loading_jobs SET status='completed', completed_at=CURRENT_TIMESTAMP "
               "WHERE shift_id=? AND status='active'", (shift['id'],))
    logger.info(f"🕐 Shift ended: {shift['id']}")
    return jsonify({'message': 'Shift ended', 'shift_id': shift['id']})

@app.route('/api/shifts/<shift_id>', methods=['GET'])
@token_required
def get_shift(shift_id):
    shift = db.fetchone("SELECT * FROM shifts WHERE id=?", (shift_id,))
    if not shift:
        return jsonify({'error': 'Not found'}), 404
    jobs = db.fetchall("SELECT * FROM loading_jobs WHERE shift_id=?", (shift_id,))
    counts = db.fetchone(
        "SELECT SUM(CASE WHEN direction='IN' THEN 1 ELSE 0 END) as total_in, "
        "SUM(CASE WHEN direction='OUT' THEN 1 ELSE 0 END) as total_out "
        "FROM detections WHERE shift_id=?", (shift_id,))
    return jsonify({'shift': shift, 'jobs': jobs, 'counts': counts})


# ============================================================================
# ROUTES: LOADING JOBS
# ============================================================================

@app.route('/api/jobs', methods=['GET'])
@token_required
def list_jobs():
    status = request.args.get('status')
    if status:
        jobs = db.fetchall("SELECT * FROM loading_jobs WHERE status=? ORDER BY created_at DESC LIMIT 100", (status,))
    else:
        jobs = db.fetchall("SELECT * FROM loading_jobs ORDER BY created_at DESC LIMIT 100")
    return jsonify({'jobs': jobs})

@app.route('/api/jobs', methods=['POST'])
@token_required
def create_job():
    d = request.get_json(silent=True) or {}
    if not d.get('truck_plate') or not d.get('target_count'):
        raise ValidationError("truck_plate and target_count are required")
    # Get active shift
    shift = db.fetchone("SELECT id FROM shifts WHERE ended_at IS NULL ORDER BY started_at DESC LIMIT 1")
    job_id = str(uuid.uuid4())
    db.insert('loading_jobs', {
        'id': job_id,
        'bay_id': d.get('bay_id'),
        'truck_plate': d['truck_plate'].upper().strip(),
        'product_name': d.get('product_name', 'Sugar Bag'),
        'target_count': int(d['target_count']),
        'direction': d.get('direction', 'OUT'),
        'shift_id': shift['id'] if shift else None,
        'operator_name': d.get('operator_name', ''),
        'notes': d.get('notes', ''),
        'status': 'pending',
    })
    logger.info(f"📋 Loading job created: {job_id} | Truck: {d['truck_plate']} | Target: {d['target_count']}")
    return jsonify({'id': job_id, 'message': 'Job created'}), 201

@app.route('/api/jobs/<job_id>', methods=['GET'])
@token_required
def get_job(job_id):
    job = db.fetchone("SELECT * FROM loading_jobs WHERE id=?", (job_id,))
    if not job:
        return jsonify({'error': 'Not found'}), 404
    detections = db.fetchall(
        "SELECT * FROM detections WHERE job_id=? ORDER BY detected_at DESC LIMIT 50", (job_id,))
    return jsonify({'job': job, 'detections': detections})

@app.route('/api/jobs/<job_id>/start', methods=['POST'])
@token_required
def start_job(job_id):
    job = db.fetchone("SELECT * FROM loading_jobs WHERE id=?", (job_id,))
    if not job:
        return jsonify({'error': 'Not found'}), 404
    db.execute("UPDATE loading_jobs SET status='active', started_at=CURRENT_TIMESTAMP WHERE id=?", (job_id,))
    # Attach job to bay camera if bay is set
    if job['bay_id']:
        cam = camera_manager.cameras.get(job['bay_id'])
        if cam:
            cam.active_job_id = job_id
    logger.info(f"▶️  Job started: {job_id}")
    return jsonify({'message': 'Job started'})

@app.route('/api/jobs/<job_id>/complete', methods=['PATCH'])
@token_required
def complete_job(job_id):
    job = db.fetchone("SELECT * FROM loading_jobs WHERE id=?", (job_id,))
    if not job:
        return jsonify({'error': 'Not found'}), 404
    db.execute("UPDATE loading_jobs SET status='completed', completed_at=CURRENT_TIMESTAMP WHERE id=?", (job_id,))
    if job['bay_id']:
        cam = camera_manager.cameras.get(job['bay_id'])
        if cam and getattr(cam, 'active_job_id', None) == job_id:
            cam.active_job_id = None
    create_alert('job_completed', f"Job completed: Truck {job['truck_plate']} — {job['loaded_count']}/{job['target_count']} bags", 'info')
    logger.info(f"✅ Job completed: {job_id}")
    return jsonify({'message': 'Job completed'})

@app.route('/api/jobs/<job_id>/cancel', methods=['PATCH'])
@token_required
def cancel_job(job_id):
    db.execute("UPDATE loading_jobs SET status='cancelled' WHERE id=?", (job_id,))
    return jsonify({'message': 'Job cancelled'})


def _increment_job_count(job_id: str, direction: str):
    """Called from camera tracking when a bag crosses the line with an active job."""
    if not job_id:
        return
    job = db.fetchone("SELECT * FROM loading_jobs WHERE id=? AND status='active'", (job_id,))
    if not job:
        return
    new_count = job['loaded_count'] + 1
    db.execute("UPDATE loading_jobs SET loaded_count=? WHERE id=?", (new_count, job_id))
    target = job['target_count']
    if new_count >= target:
        db.execute("UPDATE loading_jobs SET status='completed', completed_at=CURRENT_TIMESTAMP WHERE id=?", (job_id,))
        create_alert('job_complete', f"✅ Truck {job['truck_plate']}: {new_count}/{target} bags loaded — COMPLETE", 'success')
        logger.info(f"🎯 Job {job_id} reached target ({new_count}/{target})")
    elif new_count == target - 10:
        create_alert('job_near_complete', f"⚠️ Truck {job['truck_plate']}: {new_count}/{target} bags — 10 remaining", 'warning')


# ============================================================================
# ROUTES: REPORTS  (PDF + Excel)
# ============================================================================

def _build_daily_data(date_str: str) -> dict:
    """Collect all data for a given date (YYYY-MM-DD) to use in reports."""
    detections = db.fetchall(
        "SELECT * FROM detections WHERE DATE(detected_at)=? ORDER BY detected_at", (date_str,))
    jobs = db.fetchall(
        "SELECT * FROM loading_jobs WHERE DATE(created_at)=? ORDER BY created_at", (date_str,))
    shifts = db.fetchall(
        "SELECT * FROM shifts WHERE DATE(started_at)=? ORDER BY started_at", (date_str,))
    totals = db.fetchone(
        "SELECT SUM(CASE WHEN direction='IN' THEN 1 ELSE 0 END) as total_in, "
        "SUM(CASE WHEN direction='OUT' THEN 1 ELSE 0 END) as total_out, "
        "COUNT(*) as total FROM detections WHERE DATE(detected_at)=?", (date_str,))
    return {'date': date_str, 'detections': detections, 'jobs': jobs,
            'shifts': shifts, 'totals': totals or {}}

@app.route('/api/reports/daily')
@token_required
def daily_report_json():
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    return jsonify(_build_daily_data(date_str))

@app.route('/api/reports/pdf')
@token_required
def daily_report_pdf():
    """Generate and stream a PDF daily report."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        return jsonify({'error': 'reportlab not installed'}), 500

    import io
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    data = _build_daily_data(date_str)
    totals = data['totals']

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph(f"<b>AI CCTV Daily Report</b>", styles['Title']))
    story.append(Paragraph(f"Date: {date_str}", styles['Normal']))
    story.append(Spacer(1, 0.5*cm))

    # Summary box
    summary = [
        ['Metric', 'Value'],
        ['Total Detections', str(totals.get('total', 0))],
        ['Bags IN', str(totals.get('total_in', 0))],
        ['Bags OUT', str(totals.get('total_out', 0))],
        ['Loading Jobs', str(len(data['jobs']))],
        ['Shifts', str(len(data['shifts']))],
    ]
    t = Table(summary, colWidths=[8*cm, 6*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f7ff')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f0ff')]),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # Loading Jobs table
    if data['jobs']:
        story.append(Paragraph('<b>Loading Jobs</b>', styles['Heading2']))
        job_rows = [['Truck', 'Product', 'Target', 'Loaded', 'Status', 'Operator']]
        for j in data['jobs']:
            job_rows.append([j['truck_plate'], j['product_name'], str(j['target_count']),
                             str(j['loaded_count']), j['status'], j.get('operator_name', '—')])
        jt = Table(job_rows, colWidths=[3*cm, 3*cm, 2.5*cm, 2.5*cm, 2.5*cm, 3*cm])
        jt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e1b4b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f0ff')]),
        ]))
        story.append(jt)
        story.append(Spacer(1, 0.5*cm))

    # Shifts table
    if data['shifts']:
        story.append(Paragraph('<b>Shifts</b>', styles['Heading2']))
        shift_rows = [['Name', 'Operator', 'Started', 'Ended']]
        for s in data['shifts']:
            shift_rows.append([s['name'], s.get('operator_name', '—'),
                               str(s['started_at'])[:16], str(s.get('ended_at', 'Active'))[:16]])
        st = Table(shift_rows, colWidths=[4*cm, 4*cm, 5*cm, 4*cm])
        st.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e1b4b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f0ff')]),
        ]))
        story.append(st)

    doc.build(story)
    buf.seek(0)
    return Response(buf.read(), mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename="report_{date_str}.pdf"'})

@app.route('/api/reports/excel')
@token_required
def daily_report_excel():
    """Generate and stream an Excel daily report."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl not installed'}), 500

    import io
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    data = _build_daily_data(date_str)

    wb = openpyxl.Workbook()
    # ── Summary sheet ──
    ws = wb.active
    ws.title = 'Summary'
    hdr_fill = PatternFill('solid', fgColor='7C3AED')
    hdr_font = Font(color='FFFFFF', bold=True)
    ws.append(['AI CCTV Daily Report', date_str])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(['Metric', 'Value'])
    for cell in ws[3]: cell.fill = hdr_fill; cell.font = hdr_font
    totals = data['totals']
    for row in [['Total Detections', totals.get('total', 0)],
                ['Bags IN', totals.get('total_in', 0)],
                ['Bags OUT', totals.get('total_out', 0)],
                ['Loading Jobs', len(data['jobs'])],
                ['Shifts', len(data['shifts'])]]:
        ws.append(row)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

    # ── Jobs sheet ──
    ws2 = wb.create_sheet('Loading Jobs')
    headers = ['Truck Plate', 'Product', 'Target', 'Loaded', 'Status', 'Operator', 'Bay', 'Started', 'Completed']
    ws2.append(headers)
    for cell in ws2[1]: cell.fill = hdr_fill; cell.font = hdr_font
    for j in data['jobs']:
        ws2.append([j['truck_plate'], j['product_name'], j['target_count'], j['loaded_count'],
                    j['status'], j.get('operator_name', ''), j.get('bay_id', ''),
                    str(j.get('started_at', '')), str(j.get('completed_at', ''))])
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 16

    # ── Shifts sheet ──
    ws3 = wb.create_sheet('Shifts')
    ws3.append(['Name', 'Operator', 'Started', 'Ended', 'Notes'])
    for cell in ws3[1]: cell.fill = hdr_fill; cell.font = hdr_font
    for s in data['shifts']:
        ws3.append([s['name'], s.get('operator_name', ''), str(s['started_at']),
                    str(s.get('ended_at', 'Active')), s.get('notes', '')])
    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.read(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="report_{date_str}.xlsx"'})


# ============================================================================
# ROUTES: COMPRESSION
# ============================================================================

@app.route('/api/compression/upload', methods=['POST'])
@token_required
def compress_upload():
    if 'file' not in request.files: raise ValidationError("No file")
    f = request.files['file']
    jid = str(uuid.uuid4())[:8]
    inp = Config.UPLOAD_DIR / f"{jid}_in{Path(f.filename).suffix}"
    out = Config.UPLOAD_DIR / f"{jid}_out.mp4"
    f.save(inp)
    
    job = CompressionJob(jid, inp, out, original_size=inp.stat().st_size)
    compression_jobs[jid] = job
    
    def compress():
        try:
            subprocess.run(['ffmpeg', '-y', '-i', str(inp), '-c:v', 'libx264', '-crf', '28', str(out)], capture_output=True, timeout=Config.COMPRESSION_TIMEOUT)
            if out.exists():
                job.status = JobStatus.COMPLETED
                job.compressed_size = out.stat().st_size
            else: job.status = JobStatus.FAILED
        except Exception as e:
            job.status = JobStatus.FAILED
            job.error = str(e)
    
    threading.Thread(target=compress, daemon=True).start()
    return jsonify({'job_id': jid, 'status': 'processing'}), 202

@app.route('/api/compression/status/<jid>')
@token_required
def compression_status(jid):
    job = compression_jobs.get(jid)
    if not job: raise NotFoundError("Job")
    return jsonify(job.to_dict())

@app.route('/api/compression/download/<jid>')
def compression_download(jid):
    job = compression_jobs.get(jid)
    if not job or job.status != JobStatus.COMPLETED: raise NotFoundError("File")
    return send_file(job.output_path, as_attachment=True)

# ============================================================================
# VIDEO PROCESSING - DATA STRUCTURES
# ============================================================================

from collections import deque
from typing import Set

@dataclass
class VideoJob:
    id: str
    status: str  # queued, processing, completed, failed, cancelled
    progress: int
    frames_total: int
    frames_processed: int
    offloaded_count: int
    loaded_count: int = 0
    detections_total: int = 0
    error: Optional[str] = None
    input_path: Optional[Path] = None
    output_path: Optional[Path] = None
    created_at: datetime = field(default_factory=datetime.now)
    completed_at: Optional[datetime] = None
    config: dict = field(default_factory=dict)
    
    def to_dict(self) -> dict:
        return {
            'job_id': self.id,
            'status': self.status,
            'progress': self.progress,
            'frames_total': self.frames_total,
            'frames_processed': self.frames_processed,
            'offloaded_count': self.offloaded_count,
            'loaded_count': self.loaded_count,
            'detections_total': self.detections_total,
            'error': self.error,
            'has_output_video': self.output_path is not None and self.output_path.exists() if self.output_path else False,
            'config': self.config
        }

video_processing_jobs: Dict[str, VideoJob] = {}
video_processing_lock = threading.Lock()
live_export_jobs: Dict[str, VideoJob] = {}
live_export_lock = threading.Lock()

def cleanup_old_video_jobs(max_age_hours: int = 6):
    """Clean up old video processing jobs"""
    cutoff = datetime.now() - timedelta(hours=max_age_hours)
    with video_processing_lock:
        to_delete = [jid for jid, job in video_processing_jobs.items() 
                     if job.created_at < cutoff]
        for jid in to_delete:
            job = video_processing_jobs[jid]
            if job.input_path and job.input_path.exists():
                try: job.input_path.unlink()
                except: pass
            if job.output_path and job.output_path.exists():
                try: job.output_path.unlink()
                except: pass
            del video_processing_jobs[jid]
    cleanup_old_live_export_jobs(max_age_hours=max_age_hours)

def cleanup_old_live_export_jobs(max_age_hours: int = 6):
    cutoff = datetime.now() - timedelta(hours=max_age_hours)
    with live_export_lock:
        to_delete = [jid for jid, job in live_export_jobs.items() if job.created_at < cutoff]
        for jid in to_delete:
            job = live_export_jobs[jid]
            for path in (job.input_path, job.output_path):
                if path and path.exists():
                    try: path.unlink()
                    except: pass
            del live_export_jobs[jid]

def _parse_live_export_time(value: str, field: str) -> datetime:
    if not value:
        raise ValidationError(f"{field} is required")
    raw = str(value).strip()
    if raw.endswith('Z'):
        raw = raw[:-1] + '+00:00'
    try:
        dt = datetime.fromisoformat(raw)
    except ValueError:
        raise ValidationError(f"{field} must be ISO datetime, for example 2026-08-18T12:30:00")
    if dt.tzinfo is not None:
        dt = dt.astimezone().replace(tzinfo=None)
    return dt

def _iter_live_recording_chunks(start_ms: int, end_ms: int, camera_id: str = None, source: str = None):
    chunks = []
    for meta_path in sorted(Config.LIVE_RECORDING_DIR.glob('live_*.jsonl')):
        video_path = meta_path.with_suffix('.mp4')
        if not video_path.exists():
            continue
        first_ms = last_ms = None
        matched_source = False
        try:
            with meta_path.open('r', encoding='utf-8') as fh:
                for line in fh:
                    if not line.strip():
                        continue
                    item = json.loads(line)
                    if camera_id and item.get('camera_id') != camera_id:
                        continue
                    if source and str(item.get('source', '')) != str(source):
                        continue
                    matched_source = True
                    ts = int(item.get('epoch_ms', 0))
                    if first_ms is None:
                        first_ms = ts
                    last_ms = ts
        except Exception as e:
            logger.warning(f"Skipping corrupt live metadata {meta_path}: {e}")
            continue
        if first_ms is None or last_ms is None:
            continue
        if (camera_id or source) and not matched_source:
            continue
        if last_ms >= start_ms and first_ms <= end_ms:
            chunks.append((video_path, meta_path))
    return chunks

def _draw_live_export_overlay(frame: np.ndarray, meta: dict):
    h, w = frame.shape[:2]
    line_pos = float(meta.get('line_position', 0.45))
    line_x = int(max(0.0, min(1.0, line_pos)) * w)
    cv2.line(frame, (line_x, 0), (line_x, h), (255, 0, 0), 2)
    cv2.putText(frame, "TRACKING LINE", (min(line_x + 10, max(10, w - 180)), 30),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 0, 0), 2)

    for det in meta.get('detections') or []:
        try:
            x1, y1, x2, y2 = [int(v) for v in det.get('bbox', [])[:4]]
        except Exception:
            continue
        label = det.get('label', 'object')
        conf = float(det.get('confidence', 0))
        cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
        cv2.putText(frame, f"{label}: {conf:.2f}", (x1, max(18, y1 - 8)),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)

    for trk in meta.get('tracks') or []:
        try:
            cx = int(float(trk.get('cx', 0)))
            cy = int(float(trk.get('cy', 0)))
            oid = int(trk.get('id', 0))
        except Exception:
            continue
        cv2.circle(frame, (cx, cy), 4, (0, 0, 255), -1)
        cv2.putText(frame, f"ID {oid}", (cx - 10, cy - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 1)

    cv2.putText(frame, f"IN:  {int(meta.get('session_in', 0))}", (10, h - 44),
                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (50, 220, 50), 2)
    cv2.putText(frame, f"OUT: {int(meta.get('session_out', 0))}", (10, h - 14),
                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (50, 80, 255), 2)
    ts = str(meta.get('ts', ''))
    if ts:
        cv2.putText(frame, ts.replace('T', ' '), (10, 28),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 255), 2)

def _compress_video_if_possible(input_path: Path, output_path: Path):
    try:
        result = subprocess.run(
            ['ffmpeg', '-y', '-i', str(input_path), '-c:v', 'libx264', '-preset', 'veryfast',
             '-crf', '28', '-movflags', '+faststart', str(output_path)],
            capture_output=True,
            timeout=Config.COMPRESSION_TIMEOUT,
        )
        if result.returncode == 0 and output_path.exists() and output_path.stat().st_size > 0:
            return output_path
        logger.warning(f"ffmpeg live export compression failed: {result.stderr[-500:] if result.stderr else b''}")
    except Exception as e:
        logger.warning(f"ffmpeg live export compression unavailable: {e}")
    return input_path

def _build_live_export(job: VideoJob, start_dt: datetime, end_dt: datetime):
    start_ms = int(start_dt.timestamp() * 1000)
    end_ms = int(end_dt.timestamp() * 1000)
    camera_id = job.config.get('camera_id') if job.config else None
    source = job.config.get('source') if job.config else None
    raw_output = Config.UPLOAD_DIR / f"{job.id}_live_raw.mp4"
    final_output = Config.UPLOAD_DIR / f"{job.id}_live_export.mp4"
    writer = None
    frames_written = 0
    frames_seen = 0
    output_size = None
    chunks = _iter_live_recording_chunks(start_ms, end_ms, camera_id=camera_id, source=source)
    job.frames_total = max(1, len(chunks))
    try:
        if not chunks:
            job.status = 'failed'
            job.error = 'No live recording chunks found for this time range'
            return
        job.status = 'processing'
        for chunk_index, (video_path, meta_path) in enumerate(chunks, start=1):
            cap = cv2.VideoCapture(str(video_path))
            try:
                with meta_path.open('r', encoding='utf-8') as fh:
                    for line in fh:
                        if not line.strip():
                            continue
                        ret, frame = cap.read()
                        if not ret:
                            break
                        frames_seen += 1
                        meta = json.loads(line)
                        ts_ms = int(meta.get('epoch_ms', 0))
                        if camera_id and meta.get('camera_id') != camera_id:
                            continue
                        if source and str(meta.get('source', '')) != str(source):
                            continue
                        if ts_ms < start_ms or ts_ms > end_ms:
                            continue
                        if writer is None:
                            h, w = frame.shape[:2]
                            output_size = (w, h)
                            fourcc = cv2.VideoWriter_fourcc(*'mp4v')
                            writer = cv2.VideoWriter(str(raw_output), fourcc, Config.LIVE_RECORDING_FPS, output_size)
                            if not writer.isOpened():
                                job.status = 'failed'
                                job.error = 'Could not create live export video'
                                return
                        elif output_size and (frame.shape[1], frame.shape[0]) != output_size:
                            frame = cv2.resize(frame, output_size, interpolation=cv2.INTER_LINEAR)
                        _draw_live_export_overlay(frame, meta)
                        writer.write(frame)
                        frames_written += 1
                        job.frames_processed = frames_written
                        job.detections_total += len(meta.get('detections') or [])
            finally:
                cap.release()
            job.progress = min(95, int((chunk_index / max(1, len(chunks))) * 95))

        if writer:
            writer.release()
            writer = None
        if frames_written <= 0:
            job.status = 'failed'
            job.error = 'No recorded frames found inside this time range'
            return

        chosen = _compress_video_if_possible(raw_output, final_output)
        job.output_path = chosen
        job.input_path = raw_output if chosen != raw_output else None
        if chosen != raw_output and raw_output.exists():
            try: raw_output.unlink()
            except: pass
        job.status = 'completed'
        job.progress = 100
        job.completed_at = datetime.now()
        logger.info(f"Live export {job.id} completed: {frames_written} frames from {len(chunks)} chunks")
    except Exception as e:
        job.status = 'failed'
        job.error = str(e)
        logger.error(f"Live export {job.id} failed: {e}")
    finally:
        if writer:
            writer.release()

# ============================================================================
# ROUTES: LIVE RECORDING EXPORT
# ============================================================================

@app.route('/api/live_recordings/status')
@token_required
def live_recordings_status():
    chunks = _iter_live_recording_chunks(0, int(time.time() * 1000))
    oldest = newest = None
    total_bytes = 0
    sources = {}
    for video_path, meta_path in chunks:
        try:
            total_bytes += video_path.stat().st_size + meta_path.stat().st_size
            with meta_path.open('r', encoding='utf-8') as fh:
                for line in fh:
                    if not line.strip():
                        continue
                    item = json.loads(line)
                    ts = item.get('ts')
                    sid = item.get('camera_id') or item.get('source') or 'unknown'
                    source_info = sources.setdefault(sid, {
                        'camera_id': item.get('camera_id'),
                        'source': item.get('source'),
                        'source_kind': item.get('source_kind'),
                        'oldest': ts,
                        'newest': ts,
                        'frames': 0,
                    })
                    source_info['newest'] = ts
                    source_info['frames'] += 1
                    if oldest is None:
                        oldest = ts
                    newest = ts
        except Exception:
            continue
    return jsonify({
        'enabled': bool(getattr(camera, 'recording_enabled', True)) if camera else True,
        'active': camera is not None,
        'source': str(camera.source) if camera else None,
        'source_kind': camera.source_kind if camera else None,
        'fps': Config.LIVE_RECORDING_FPS,
        'chunk_seconds': Config.LIVE_RECORDING_CHUNK_SECONDS,
        'retention_hours': Config.LIVE_RECORDING_RETENTION_HOURS,
        'chunks': len(chunks),
        'oldest': oldest,
        'newest': newest,
        'bytes': total_bytes,
        'sources': list(sources.values()),
    })

@app.route('/api/live_recordings/export', methods=['POST'])
@token_required
def live_recordings_export():
    data = request.get_json(silent=True) or {}
    if data.get('start') or data.get('end'):
        start_dt = _parse_live_export_time(data.get('start'), 'start')
        end_dt = _parse_live_export_time(data.get('end'), 'end')
    else:
        end_dt = datetime.now()
        start_dt = end_dt - timedelta(hours=1)
    if end_dt <= start_dt:
        raise ValidationError("end must be after start")
    max_hours = min(Config.LIVE_RECORDING_RETENTION_HOURS, int(os.getenv('LIVE_EXPORT_MAX_HOURS', '6')))
    if end_dt - start_dt > timedelta(hours=max_hours):
        raise ValidationError(f"range cannot exceed {max_hours} hour(s)")

    camera_id = data.get('camera_id') or data.get('bay_id')
    source = str(normalize_camera_source(data.get('source'))) if data.get('source') is not None else None
    job_id = str(uuid.uuid4())[:8]
    job_config = {
        'start': start_dt.isoformat(),
        'end': end_dt.isoformat(),
        'fps': Config.LIVE_RECORDING_FPS,
    }
    if camera_id:
        job_config['camera_id'] = camera_id
    if source:
        job_config['source'] = source
    job = VideoJob(
        id=job_id,
        status='queued',
        progress=0,
        frames_total=0,
        frames_processed=0,
        offloaded_count=0,
        output_path=Config.UPLOAD_DIR / f"{job_id}_live_export.mp4",
        config=job_config,
    )
    with live_export_lock:
        live_export_jobs[job_id] = job
    threading.Thread(target=_build_live_export, args=(job, start_dt, end_dt), daemon=True).start()
    return jsonify({'job_id': job_id, 'status': job.status, 'config': job.config}), 202

@app.route('/api/live_recordings/export/<job_id>')
@token_required
def live_recordings_export_status(job_id):
    job = live_export_jobs.get(job_id)
    if not job:
        raise NotFoundError("Live export job")
    return jsonify(job.to_dict())

@app.route('/api/live_recordings/export/<job_id>/download')
def live_recordings_export_download(job_id):
    job = live_export_jobs.get(job_id)
    if not job or job.status != 'completed' or not job.output_path or not job.output_path.exists():
        raise NotFoundError("Live export")
    return send_file(job.output_path, as_attachment=True, download_name=f'live_detected_{job_id}.mp4')

def cleanup_old_compression_jobs(max_age_hours: int = 6):
    """Clean up completed/failed compression jobs to prevent unbounded dict growth."""
    cutoff = datetime.now() - timedelta(hours=max_age_hours)
    stale = [jid for jid, job in list(compression_jobs.items())
             if job.status in (JobStatus.COMPLETED, JobStatus.FAILED)]
    for jid in stale:
        job = compression_jobs[jid]
        for path in (job.input_path, job.output_path):
            if path and path.exists():
                try: path.unlink()
                except: pass
        del compression_jobs[jid]

def save_video_detection(det_type: str, conf: float, direction: str, count: int = 1):
    """Batch-insert video detections in a single transaction."""
    if count <= 0:
        return
    rows = [{'id': str(uuid.uuid4()), 'type': det_type, 'confidence': conf,
             'direction': direction, 'inference_ms': 0}
            for _ in range(count)]
    try:
        with db.get_cursor() as cursor:
            cursor.executemany(
                "INSERT INTO detections (id, type, confidence, direction, inference_ms) "
                "VALUES (:id, :type, :confidence, :direction, :inference_ms)",
                rows
            )
    except Exception as e:
        logger.error(f"Failed to batch-save video detections: {e}")

# ============================================================================
# SIMPLE TRACKER (for version_5)
# ============================================================================

class SimpleCentroidTracker:
    """Simple centroid tracker for basic left→right counting"""
    
    def __init__(self, max_disappeared: int = 30):
        self.next_id = 0
        self.objects: Dict[int, Tuple[float, float]] = {}
        self.disappeared: Dict[int, int] = {}
        self.max_disappeared = max_disappeared
        self.crossed: set = set()
    
    def register(self, centroid: Tuple[float, float]) -> int:
        self.objects[self.next_id] = centroid
        self.disappeared[self.next_id] = 0
        self.next_id += 1
        return self.next_id - 1
    
    def deregister(self, object_id: int):
        del self.objects[object_id]
        del self.disappeared[object_id]
    
    def update(self, detections: List[Tuple[float, float]]) -> Dict[int, Tuple[float, float]]:
        if len(detections) == 0:
            for oid in list(self.disappeared.keys()):
                self.disappeared[oid] += 1
                if self.disappeared[oid] > self.max_disappeared:
                    self.deregister(oid)
            return self.objects
        
        if len(self.objects) == 0:
            for det in detections:
                self.register(det)
        else:
            object_ids = list(self.objects.keys())
            object_centroids = list(self.objects.values())
            used_detections = set()
            
            for i, oid in enumerate(object_ids):
                ox, oy = object_centroids[i]
                min_dist = float('inf')
                best_det_idx = -1
                
                for j, (dx, dy) in enumerate(detections):
                    if j in used_detections:
                        continue
                    dist = ((ox - dx) ** 2 + (oy - dy) ** 2) ** 0.5
                    if dist < min_dist and dist < 100:
                        min_dist = dist
                        best_det_idx = j
                
                if best_det_idx != -1:
                    self.objects[oid] = detections[best_det_idx]
                    self.disappeared[oid] = 0
                    used_detections.add(best_det_idx)
                else:
                    self.disappeared[oid] += 1
                    if self.disappeared[oid] > self.max_disappeared:
                        self.deregister(oid)
            
            for j, det in enumerate(detections):
                if j not in used_detections:
                    self.register(det)
        
        return self.objects

# ============================================================================
# OPTIMIZED TRACKER (for version_2)
# ============================================================================

@dataclass
class CountingConfig:
    confidence_threshold: float = 0.4
    nms_threshold: float = 0.3
    min_box_area_ratio: float = 0.005
    max_box_area_ratio: float = 0.25
    min_aspect_ratio: float = 0.3
    max_aspect_ratio: float = 3.0
    max_disappeared_frames: int = 20
    max_distance_ratio: float = 0.12
    min_track_frames: int = 3
    line_position: float = 0.45  # 45% from left
    count_zone_width: float = 0.15
    require_direction_frames: int = 5
    min_travel_distance: float = 0.1
    cooldown_frames: int = 15
    duplicate_cooldown_frames: int = 30
    duplicate_y_tolerance: float = 0.20
    line_clear_frames: int = 3
    roi_x1: float = 0.0
    roi_x2: float = 1.0
    roi_y1: float = 0.0
    roi_y2: float = 1.0

@dataclass
class TrackedObject:
    id: int
    label: str
    confidence: float
    bbox: List[int]
    centroid: Tuple[float, float]
    position_history: deque = field(default_factory=lambda: deque(maxlen=30))
    confidence_history: deque = field(default_factory=lambda: deque(maxlen=10))
    frames_tracked: int = 0
    frames_disappeared: int = 0
    first_seen_frame: int = 0
    last_seen_frame: int = 0
    has_crossed: bool = False
    cross_direction: Optional[str] = None
    cross_frame: Optional[int] = None
    entry_position: Optional[float] = None
    
    def add_position(self, cx: float, cy: float, frame_num: int):
        self.position_history.append((cx, cy, frame_num))
        self.last_seen_frame = frame_num
        self.frames_tracked += 1
        self.frames_disappeared = 0
    
    def add_confidence(self, conf: float):
        self.confidence_history.append(conf)
    
    def get_average_confidence(self) -> float:
        if not self.confidence_history: return self.confidence
        return sum(self.confidence_history) / len(self.confidence_history)
    
    def get_velocity(self) -> Tuple[float, float]:
        if len(self.position_history) < 2: return (0.0, 0.0)
        recent = list(self.position_history)[-5:]
        if len(recent) < 2: return (0.0, 0.0)
        dx = recent[-1][0] - recent[0][0]
        dy = recent[-1][1] - recent[0][1]
        frames = recent[-1][2] - recent[0][2]
        if frames == 0: return (0.0, 0.0)
        return (dx / frames, dy / frames)
    
    def get_travel_distance(self) -> float:
        if not self.position_history or self.entry_position is None: return 0.0
        return abs(self.position_history[-1][0] - self.entry_position)
    
    def is_moving_consistently(self, required_frames: int = 5) -> bool:
        if len(self.position_history) < required_frames: return False
        recent = list(self.position_history)[-required_frames:]
        directions = []
        for i in range(1, len(recent)):
            dx = recent[i][0] - recent[i-1][0]
            if abs(dx) > 0.005:
                directions.append('right' if dx > 0 else 'left')
        if not directions: return False
        return len(set(directions)) == 1

class EnhancedTracker:
    def __init__(self, config: CountingConfig):
        self.config = config
        self.next_id = 0
        self.objects: Dict[int, TrackedObject] = {}
        self.frame_count = 0
        self.counted_ids: Set[int] = set()
        self.count_cooldowns: Dict[int, int] = {}
        self.recent_crossings: deque = deque(maxlen=50)
        self.line_latched_direction: Optional[str] = None
        self.line_clear_streak = config.line_clear_frames
        self.total_in = 0
        self.total_out = 0
        self.rejected_detections = 0
    
    def _filter_detection(self, bbox: List[int], frame_width: int, frame_height: int) -> bool:
        x1, y1, x2, y2 = bbox
        width = x2 - x1
        height = y2 - y1
        if width <= 0 or height <= 0: return False
        frame_area = frame_width * frame_height
        box_area = width * height
        area_ratio = box_area / frame_area
        if area_ratio < self.config.min_box_area_ratio: return False
        if area_ratio > self.config.max_box_area_ratio: return False
        aspect_ratio = width / height
        if aspect_ratio < self.config.min_aspect_ratio: return False
        if aspect_ratio > self.config.max_aspect_ratio: return False
        cx = (x1 + x2) / 2 / frame_width
        cy = (y1 + y2) / 2 / frame_height
        if not (self.config.roi_x1 <= cx <= self.config.roi_x2): return False
        if not (self.config.roi_y1 <= cy <= self.config.roi_y2): return False
        return True
    
    def _calculate_iou(self, box1: List[int], box2: List[int]) -> float:
        x1 = max(box1[0], box2[0])
        y1 = max(box1[1], box2[1])
        x2 = min(box1[2], box2[2])
        y2 = min(box1[3], box2[3])
        intersection = max(0, x2 - x1) * max(0, y2 - y1)
        area1 = (box1[2] - box1[0]) * (box1[3] - box1[1])
        area2 = (box2[2] - box2[0]) * (box2[3] - box2[1])
        union = area1 + area2 - intersection
        return intersection / union if union > 0 else 0
    
    def _apply_nms(self, detections: List[Dict], threshold: float) -> List[Dict]:
        if len(detections) == 0: return []
        detections = sorted(detections, key=lambda x: x['confidence'], reverse=True)
        kept = []
        for det in detections:
            dominated = False
            for kept_det in kept:
                if self._calculate_iou(det['bbox'], kept_det['bbox']) > threshold:
                    dominated = True
                    break
            if not dominated: kept.append(det)
        return kept
    
    def update(self, raw_detections: List[Dict], frame_width: int, frame_height: int) -> Dict[int, TrackedObject]:
        self.frame_count += 1
        valid_detections = []
        for det in raw_detections:
            if self._filter_detection(det['bbox'], frame_width, frame_height):
                x1, y1, x2, y2 = det['bbox']
                cx = ((x1 + x2) / 2) / frame_width
                cy = ((y1 + y2) / 2) / frame_height
                valid_detections.append({**det, 'cx': cx, 'cy': cy})
            else:
                self.rejected_detections += 1
        valid_detections = self._apply_nms(valid_detections, self.config.nms_threshold)
        zone_half = self.config.count_zone_width / 2
        line_has_detection = any(
            (self.config.line_position - zone_half) <= det['cx'] <= (self.config.line_position + zone_half)
            for det in valid_detections
        )
        if line_has_detection:
            self.line_clear_streak = 0
        else:
            self.line_clear_streak += 1
            if self.line_clear_streak >= self.config.line_clear_frames:
                self.line_latched_direction = None
        
        if len(valid_detections) == 0:
            for oid in list(self.objects.keys()):
                self.objects[oid].frames_disappeared += 1
                if self.objects[oid].frames_disappeared > self.config.max_disappeared_frames:
                    del self.objects[oid]
            return self.objects
        
        if len(self.objects) == 0:
            for det in valid_detections:
                self._register_object(det)
            return self.objects
        
        object_ids = list(self.objects.keys())
        used_detections = set()
        matched_objects = set()
        matches = []
        
        for oid in object_ids:
            obj = self.objects[oid]
            vx, vy = obj.get_velocity()
            pred_cx, pred_cy = obj.centroid[0] + vx, obj.centroid[1] + vy
            for j, det in enumerate(valid_detections):
                dist = ((pred_cx - det['cx']) ** 2 + (pred_cy - det['cy']) ** 2) ** 0.5
                if dist < self.config.max_distance_ratio:
                    matches.append((dist, oid, j))
        
        matches.sort(key=lambda x: x[0])
        for dist, oid, det_idx in matches:
            if oid in matched_objects or det_idx in used_detections: continue
            det = valid_detections[det_idx]
            self._update_object(oid, det)
            matched_objects.add(oid)
            used_detections.add(det_idx)
        
        for oid in object_ids:
            if oid not in matched_objects:
                self.objects[oid].frames_disappeared += 1
                if self.objects[oid].frames_disappeared > self.config.max_disappeared_frames:
                    del self.objects[oid]
        
        for j, det in enumerate(valid_detections):
            if j not in used_detections:
                self._register_object(det)
        
        return self.objects
    
    def _register_object(self, det: Dict) -> int:
        obj = TrackedObject(
            id=self.next_id, label=det.get('label', 'sugar_bag'), confidence=det.get('confidence', 0.0),
            bbox=det['bbox'], centroid=(det['cx'], det['cy']),
            first_seen_frame=self.frame_count, last_seen_frame=self.frame_count, entry_position=det['cx']
        )
        obj.add_position(det['cx'], det['cy'], self.frame_count)
        obj.add_confidence(det.get('confidence', 0.0))
        self.objects[self.next_id] = obj
        self.next_id += 1
        return self.next_id - 1
    
    def _update_object(self, oid: int, det: Dict):
        obj = self.objects[oid]
        obj.bbox = det['bbox']
        obj.centroid = (det['cx'], det['cy'])
        obj.confidence = det.get('confidence', 0.0)
        obj.add_position(det['cx'], det['cy'], self.frame_count)
        obj.add_confidence(det.get('confidence', 0.0))
    
    def check_line_crossing(self, obj: TrackedObject) -> Optional[str]:
        if obj.has_crossed: return None
        if obj.id in self.count_cooldowns and self.frame_count < self.count_cooldowns[obj.id]: return None
        if obj.frames_tracked < self.config.min_track_frames: return None
        if len(obj.position_history) < self.config.require_direction_frames: return None
        
        positions = list(obj.position_history)
        current_x = positions[-1][0]
        line = self.config.line_position
        zone_half = self.config.count_zone_width / 2
        in_zone = (line - zone_half) <= current_x <= (line + zone_half)
        if not in_zone: return None
        
        prev_positions = positions[:-1]
        was_left = any(p[0] < line - zone_half for p in prev_positions[-5:])
        was_right = any(p[0] > line + zone_half for p in prev_positions[-5:])
        if not (was_left or was_right): return None
        
        if was_left and current_x >= line: direction = 'OUT'
        elif was_right and current_x <= line: direction = 'IN'
        else: return None
        
        if not obj.is_moving_consistently(self.config.require_direction_frames): return None
        if obj.get_travel_distance() < self.config.min_travel_distance: return None
        if obj.get_average_confidence() < self.config.confidence_threshold: return None

        if (
            self.line_latched_direction == direction
            and self.line_clear_streak < self.config.line_clear_frames
        ):
            obj.has_crossed = True
            obj.cross_direction = direction
            obj.cross_frame = self.frame_count
            self.counted_ids.add(obj.id)
            self.count_cooldowns[obj.id] = self.frame_count + self.config.cooldown_frames
            logger.info(f"Skipped duplicate {direction} crossing for track {obj.id}; count line is still occupied")
            return None

        current_y = positions[-1][1]
        for prev_direction, prev_y, prev_frame in self.recent_crossings:
            if direction != prev_direction:
                continue
            if self.frame_count - prev_frame > self.config.duplicate_cooldown_frames:
                continue
            if abs(current_y - prev_y) <= self.config.duplicate_y_tolerance:
                obj.has_crossed = True
                obj.cross_direction = direction
                obj.cross_frame = self.frame_count
                self.counted_ids.add(obj.id)
                self.count_cooldowns[obj.id] = self.frame_count + self.config.cooldown_frames
                logger.info(
                    f"Skipped duplicate {direction} crossing for track {obj.id} "
                    f"near y={current_y:.2f}"
                )
                return None
        
        obj.has_crossed = True
        obj.cross_direction = direction
        obj.cross_frame = self.frame_count
        self.counted_ids.add(obj.id)
        self.count_cooldowns[obj.id] = self.frame_count + self.config.cooldown_frames
        self.recent_crossings.append((direction, current_y, self.frame_count))
        self.line_latched_direction = direction
        self.line_clear_streak = 0
        
        if direction == 'OUT': self.total_out += 1
        else: self.total_in += 1
        
        return direction
    
    def get_counts(self) -> Dict:
        return {'offloaded': self.total_out, 'loaded': self.total_in, 'currently_tracking': len(self.objects), 'total_counted': len(self.counted_ids), 'rejected_detections': self.rejected_detections, 'frame_count': self.frame_count}

# ============================================================================
# ROUTES: VIDEO PROCESSING (SIMPLE - for version_5)
# ============================================================================

@app.route('/api/video/process', methods=['POST'])
@token_required
def process_video_simple():
    """Video processing with duplicate-safe left-to-right counting."""
    cleanup_old_video_jobs(max_age_hours=6)
    
    if 'file' not in request.files: raise ValidationError("No video file provided")
    video_file = request.files['file']
    if not video_file.filename: raise ValidationError("No filename")
    
    job_id = str(uuid.uuid4())[:8]
    input_path = Config.UPLOAD_DIR / f"{job_id}_input{Path(video_file.filename).suffix}"
    output_path = Config.UPLOAD_DIR / f"{job_id}_output.mp4"
    video_file.save(input_path)
    
    confidence = float(request.form.get('confidence', 0.5))
    process_fps = int(request.form.get('fps', 5))
    line_position = float(request.form.get('line_position', 0.45))
    save_video = request.form.get('save_video', 'true').lower() == 'true'

    config = CountingConfig()
    config.confidence_threshold = confidence
    config.line_position = line_position
    config.min_track_frames = int(request.form.get('min_track_frames', 3))
    config.min_travel_distance = float(request.form.get('min_travel', 0.08))
    config.count_zone_width = float(request.form.get('zone_width', 0.12))
    config.duplicate_cooldown_frames = max(
        config.cooldown_frames,
        int(float(request.form.get('duplicate_cooldown_sec', 3.0)) * max(1, process_fps)),
    )
    
    job = VideoJob(
        id=job_id,
        status='processing',
        progress=0,
        frames_total=0,
        frames_processed=0,
        offloaded_count=0,
        input_path=input_path,
        output_path=output_path if save_video else None,
    )
    job.config = {
        'confidence': confidence,
        'fps': process_fps,
        'line_position': line_position,
        'tracker': 'enhanced',
        'duplicate_cooldown_sec': round(config.duplicate_cooldown_frames / max(1, process_fps), 2),
    }
    
    with video_processing_lock:
        video_processing_jobs[job_id] = job
    
    def process():
        tracker = EnhancedTracker(config)
        cap = None
        writer = None
        
        try:
            cap = cv2.VideoCapture(str(input_path))
            if not cap.isOpened():
                job.status = 'failed'
                job.error = 'Could not open video'
                return
            
            total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            video_fps = cap.get(cv2.CAP_PROP_FPS) or 30
            frame_skip = max(1, int(video_fps / process_fps))
            frame_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            
            job.frames_total = max(1, total_frames // frame_skip)
            if save_video:
                fourcc = cv2.VideoWriter_fourcc(*'mp4v')
                writer = cv2.VideoWriter(str(output_path), fourcc, process_fps, (frame_width, frame_height))
                if not writer.isOpened():
                    job.status = 'failed'
                    job.error = 'Could not create output video'
                    return

            frame_idx = 0
            processed = 0
            
            while True:
                ret, frame = cap.read()
                if not ret: break
                frame_idx += 1
                if frame_idx % frame_skip != 0: continue
                
                processed += 1
                job.frames_processed = processed
                job.progress = min(99, int((processed / max(1, job.frames_total)) * 100))
                
                _, detections = ml_service.detect(frame, confidence=max(0.01, confidence * 0.8), draw=False)
                job.detections_total += len(detections)
                
                det_list = [{'bbox': d.bbox, 'label': d.label, 'confidence': d.confidence} for d in detections]
                tracked_objects = tracker.update(det_list, frame_width, frame_height)
                line_x = int(frame_width * line_position)
                
                for oid, obj in tracked_objects.items():
                    cross_dir = tracker.check_line_crossing(obj)
                    if cross_dir == 'OUT':
                        job.offloaded_count = tracker.total_out
                        logger.info(f"🎒 [Video] Bag {oid} offloaded! Total: {tracker.total_out}")

                if writer:
                    cv2.line(frame, (line_x, 0), (line_x, frame_height), (0, 0, 255), 3)
                    cv2.putText(frame, "COUNT LINE", (min(line_x + 10, frame_width - 170), 32),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)

                    for oid, obj in tracked_objects.items():
                        x1, y1, x2, y2 = obj.bbox
                        cx, cy = obj.centroid
                        color = (0, 165, 255) if obj.has_crossed else (0, 255, 0)
                        cv2.rectangle(frame, (x1, y1), (x2, y2), color, 3 if obj.has_crossed else 2)
                        cv2.putText(frame, f"{obj.label} ID:{oid}", (x1, max(20, y1 - 8)),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)
                        cv2.circle(frame, (int(cx * frame_width), int(cy * frame_height)), 4, color, -1)
                        cv2.putText(frame, f"ID {oid}", (int(cx * frame_width) - 10, int(cy * frame_height) - 10),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 1)

                    cv2.putText(frame, f"OFFLOADED: {tracker.total_out}", (10, 36),
                                cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 0, 255), 2)
                    writer.write(frame)
            
            cap.release()
            if writer: writer.release()
            job.status = 'completed'
            job.progress = 100
            job.offloaded_count = tracker.total_out
            job.loaded_count = tracker.total_in
            
            if job.offloaded_count > 0:
                save_video_detection('sugar_bag', 1.0, 'OUT', job.offloaded_count)
                update_inventory_from_detection('Sugar Bag', 'OUT', job.offloaded_count)
            
            logger.info(f"✅ [Video] Processed: OUT={job.offloaded_count}, IN={job.loaded_count}, frames={processed}")
            if input_path.exists(): input_path.unlink()
            
        except Exception as e:
            job.status = 'failed'
            job.error = str(e)
            logger.error(f"❌ [Simple] Video processing failed: {e}")
            if cap: cap.release()
            if writer: writer.release()
    
    threading.Thread(target=process, daemon=True).start()
    return jsonify({'job_id': job_id, 'status': 'processing', 'message': 'Simple processing started', 'config': job.config}), 202

# ============================================================================
# ROUTES: VIDEO PROCESSING (OPTIMIZED - for version_2)
# ============================================================================

@app.route('/api/video/process_optimized', methods=['POST'])
@token_required
def process_video_optimized():
    """Optimized video processing with enhanced tracking (for version_2)"""
    cleanup_old_video_jobs(max_age_hours=6)
    
    if 'file' not in request.files: raise ValidationError("No video file provided")
    video_file = request.files['file']
    if not video_file.filename: raise ValidationError("No filename")
    
    job_id = str(uuid.uuid4())[:8]
    file_ext = Path(video_file.filename).suffix.lower()
    input_path = Config.UPLOAD_DIR / f"{job_id}_input{file_ext}"
    output_path = Config.UPLOAD_DIR / f"{job_id}_output.mp4"
    video_file.save(input_path)
    
    config = CountingConfig()
    config.confidence_threshold = float(request.form.get('confidence', 0.4))
    config.line_position = float(request.form.get('line_position', 0.45))
    config.min_track_frames = int(request.form.get('min_track_frames', 3))
    config.min_travel_distance = float(request.form.get('min_travel', 0.1))
    config.count_zone_width = float(request.form.get('zone_width', 0.15))
    
    roi_str = request.form.get('roi', '')
    if roi_str:
        try:
            x1, y1, x2, y2 = map(float, roi_str.split(','))
            config.roi_x1, config.roi_y1 = max(0, x1), max(0, y1)
            config.roi_x2, config.roi_y2 = min(1, x2), min(1, y2)
        except: pass
    
    process_fps = int(request.form.get('fps', 10))
    direction_mode = request.form.get('direction', 'left_to_right')
    save_video = request.form.get('save_video', 'true').lower() == 'true'
    config.duplicate_cooldown_frames = max(
        config.cooldown_frames,
        int(float(request.form.get('duplicate_cooldown_sec', 3.0)) * max(1, process_fps)),
    )
    
    job = VideoJob(id=job_id, status='processing', progress=0, frames_total=0, frames_processed=0, offloaded_count=0, input_path=input_path, output_path=output_path if save_video else None)
    job.config = {
        'confidence': config.confidence_threshold,
        'fps': process_fps,
        'line_position': config.line_position,
        'direction': direction_mode,
        'tracker': 'optimized',
        'duplicate_cooldown_sec': round(config.duplicate_cooldown_frames / max(1, process_fps), 2),
    }
    
    with video_processing_lock:
        video_processing_jobs[job_id] = job
    
    def process():
        tracker = EnhancedTracker(config)
        cap = None
        writer = None
        
        try:
            cap = cv2.VideoCapture(str(input_path))
            if not cap.isOpened():
                job.status = 'failed'
                job.error = 'Could not open video'
                return
            
            total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            video_fps = cap.get(cv2.CAP_PROP_FPS) or 30
            frame_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            frame_skip = max(1, int(video_fps / process_fps))
            job.frames_total = max(1, total_frames // frame_skip)
            
            if save_video:
                fourcc = cv2.VideoWriter_fourcc(*'mp4v')
                writer = cv2.VideoWriter(str(output_path), fourcc, process_fps, (frame_width, frame_height))
            
            frame_idx = 0
            processed = 0
            
            while True:
                ret, frame = cap.read()
                if not ret: break
                frame_idx += 1
                if frame_idx % frame_skip != 0: continue
                
                processed += 1
                job.frames_processed = processed
                job.progress = min(99, int((processed / job.frames_total) * 100))
                
                _, detections = ml_service.detect(frame, confidence=config.confidence_threshold * 0.8, draw=False)
                job.detections_total += len(detections)
                
                det_list = [{'bbox': d.bbox, 'label': d.label, 'confidence': d.confidence} for d in detections]
                tracked = tracker.update(det_list, frame_width, frame_height)
                
                for oid, obj in tracked.items():
                    cross_dir = tracker.check_line_crossing(obj)
                    if cross_dir:
                        should_count = (direction_mode == 'both' or (direction_mode == 'left_to_right' and cross_dir == 'OUT') or (direction_mode == 'right_to_left' and cross_dir == 'IN'))
                        if should_count:
                            if cross_dir == 'OUT':
                                job.offloaded_count = tracker.total_out
                                logger.info(f"📤 [Optimized] Bag OFFLOADED! Total: {tracker.total_out}")
                            else:
                                job.loaded_count = tracker.total_in
                
                if writer:
                    # Draw bounding boxes and line
                    line_x = int(frame_width * config.line_position)
                    cv2.line(frame, (line_x, 0), (line_x, frame_height), (0, 0, 255), 2)
                    for oid, obj in tracked.items():
                        x1, y1, x2, y2 = obj.bbox
                        color = (0, 255, 0) if not obj.has_crossed else (0, 165, 255)
                        cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                        cv2.putText(frame, f"ID:{oid}", (x1, y1-5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)
                    cv2.putText(frame, f"OFFLOADED: {tracker.total_out}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                    writer.write(frame)
            
            cap.release()
            if writer: writer.release()
            
            job.status = 'completed'
            job.progress = 100
            counts = tracker.get_counts()
            job.offloaded_count = counts['offloaded']
            job.loaded_count = counts['loaded']
            
            if job.offloaded_count > 0:
                save_video_detection('sugar_bag', 1.0, 'OUT', job.offloaded_count)
                update_inventory_from_detection('Sugar Bag', 'OUT', job.offloaded_count)
            
            logger.info(f"✅ [Optimized] Complete: OUT={job.offloaded_count}, rejected={counts['rejected_detections']}")
            if input_path.exists(): input_path.unlink()
            
        except Exception as e:
            job.status = 'failed'
            job.error = str(e)
            logger.error(f"❌ [Optimized] Failed: {e}")
            if cap: cap.release()
            if writer: writer.release()
    
    threading.Thread(target=process, daemon=True).start()
    return jsonify({'job_id': job_id, 'status': 'processing', 'message': 'Optimized processing started', 'config': job.config}), 202

# ============================================================================
# ROUTES: VIDEO STATUS & DOWNLOAD
# ============================================================================

@app.route('/api/video/status/<job_id>')
@token_required
def video_status(job_id):
    job = video_processing_jobs.get(job_id)
    if not job: raise NotFoundError("Job")
    return jsonify(job.to_dict())

@app.route('/api/video/download/<job_id>')
def video_download(job_id):
    job = video_processing_jobs.get(job_id)
    if not job or job.status != 'completed' or not job.output_path or not job.output_path.exists():
        raise NotFoundError("Video")
    return send_file(job.output_path, as_attachment=True, download_name=f'detected_{job_id}.mp4')

@app.route('/api/video/cancel/<job_id>', methods=['POST'])
@token_required
def video_cancel(job_id):
    job = video_processing_jobs.get(job_id)
    if not job: raise NotFoundError("Job")
    job.status = 'cancelled'
    return jsonify({'message': 'Job cancelled'})

# ============================================================================
# ROUTES: ADMIN
# ============================================================================

@app.route('/api/reset', methods=['POST'])
@admin_required
def reset_data():
    db.delete('detections')
    db.update('inventory', {'count_in': 0, 'count_out': 0, 'current_stock': 0}, '1=1', ())
    db.delete('alerts')
    return jsonify({'status': 'reset'})

# ============================================================================
# STARTUP
# ============================================================================

@app.route('/api/video/eval_all', methods=['POST'])
def eval_all_models():
    """Single endpoint to evaluate uploaded video across all models with old & new methods."""
    if 'file' not in request.files:
        return jsonify({'error': 'No video file provided'}), 400
    video_file = request.files['file']
    line_pos = float(request.form.get('line_position', 0.5))
    conf = float(request.form.get('confidence', 0.4))
    sample_fps = int(request.form.get('fps', 10))

    temp_id = str(uuid.uuid4())[:8]
    temp_path = Config.UPLOAD_DIR / f"eval_{temp_id}.mp4"
    video_file.save(temp_path)

    model_files = [
        ('best', 'best.pt'),
        ('best_dec20', 'best_dec20.pt'),
        ('sugar_bag_final', 'sugar_bag_final.pt'),
        ('sugar_bag_finetuned', 'sugar_bag_finetuned.pt'),
        ('sugar_bag_improved', 'sugar_bag_improved.pt'),
    ]

    results = []

    try:
        from ultralytics import YOLO
        import cv2

        for name, fname in model_files:
            mpath = Config.MODEL_DIR / fname
            if not mpath.exists():
                continue

            try:
                model = YOLO(str(mpath))
            except Exception as e:
                continue

            # OLD METHOD
            cap = cv2.VideoCapture(str(temp_path))
            v_fps = cap.get(cv2.CAP_PROP_FPS) or 30
            f_skip = max(1, int(v_fps / sample_fps))
            fw = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            fh = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

            tracker_old = SimpleCentroidTracker()
            prev_pos = {}
            count_old = 0
            dets_old = 0
            fi = 0

            while True:
                ret, frame = cap.read()
                if not ret: break
                fi += 1
                if fi % f_skip != 0: continue

                res = model(frame, conf=conf, verbose=False)
                cents = []
                for r in res:
                    for box in r.boxes:
                        x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
                        cents.append(((x1 + x2) / 2, (y1 + y2) / 2))
                        dets_old += 1

                tracked = tracker_old.update(cents)
                for oid, (cx, cy) in tracked.items():
                    nx = cx / fw
                    if oid in prev_pos:
                        px = prev_pos[oid]
                        if px < line_pos <= nx and oid not in tracker_old.crossed:
                            tracker_old.crossed.add(oid)
                            count_old += 1
                    prev_pos[oid] = nx
            cap.release()

            # NEW METHOD
            cap = cv2.VideoCapture(str(temp_path))
            tracker_new = EnhancedTracker(CountingConfig(
                line_position=line_pos,
                confidence_threshold=conf,
                duplicate_cooldown_frames=max(30, int(sample_fps * 3)),
            ))
            dets_new = 0
            fi = 0

            while True:
                ret, frame = cap.read()
                if not ret: break
                fi += 1
                if fi % f_skip != 0: continue

                res = model(frame, conf=conf, verbose=False)
                raw = []
                for r in res:
                    for box in r.boxes:
                        x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
                        raw.append({'bbox': [x1, y1, x2, y2], 'label': 'bag', 'confidence': float(box.conf[0])})
                        dets_new += 1

                tracked = tracker_new.update(raw, fw, fh)
                for obj in tracked.values():
                    cross_dir = tracker_new.check_line_crossing(obj)

            cap.release()

            results.append({
                'model': name,
                'old_count': count_old,
                'old_dets': dets_old,
                'new_count': tracker_new.total_out,
                'new_dets': dets_new,
                'diff': tracker_new.total_out - count_old
            })

            del model

    finally:
        if temp_path.exists():
            temp_path.unlink()

    return jsonify({'results': results, 'line_position': line_pos})

# ============================================================================
# STATIC FRONTEND SERVING
# Serves the built React app so the entire stack runs on one port.
# API routes registered above take priority; everything else goes to index.html
# for React Router's client-side navigation.
# ============================================================================

if 'serve_assets' not in app.view_functions:
    @app.route('/assets/<path:filename>')
    def serve_assets(filename):
        """Serve Vite-built JS/CSS/image assets."""
        assets_dir = FRONTEND_DIST / 'assets'
        if assets_dir.exists():
            return send_from_directory(assets_dir, filename)
        return jsonify({'error': 'Frontend not built. Run: npm run build'}), 404


if 'serve_frontend' not in app.view_functions:
    @app.route('/', defaults={'path': ''})
    @app.route('/<path:path>')
    def serve_frontend(path):
        """Catch-all: serve React index.html for any non-API path (SPA routing)."""
        if path.startswith('api/'):
            return jsonify({'error': 'API route not found'}), 404
        if FRONTEND_DIST.exists():
            file = FRONTEND_DIST / path
            if path and file.exists() and file.is_file():
                return send_from_directory(FRONTEND_DIST, path)
            return send_from_directory(FRONTEND_DIST, 'index.html')
        return (
            '<h2>Frontend not built</h2>'
            '<p>Run: <code>cd frontends/version_5 && npm run build</code></p>'
            '<p>Then restart the backend.</p>'
        ), 200

def _maintenance_loop():
    """Background thread: periodic housekeeping every 5 minutes.
    Keeps hot-path request handlers free of expensive cleanup work.
    """
    while True:
        time.sleep(300)  # 5 minutes
        for label, fn in [
            ('low_stock_alerts', check_low_stock_alerts),
            ('rate_limiter_cleanup', rate_limiter.cleanup),
            ('compression_cleanup', cleanup_old_compression_jobs),
            ('video_job_cleanup', cleanup_old_video_jobs)
        ]:
            try:
                fn()
            except Exception as e:
                logger.error(f"Maintenance error [{label}]: {e}")

def initialize():
    print("""
    ╔══════════════════════════════════════════════════════════════╗
    ║           AI CCTV Edge (Windows ProLiant) v4.1               ║
    ╚══════════════════════════════════════════════════════════════╝
    """)
    db.init_schema()
    db.seed_data()
    if Config.MODEL_DIR.exists():
        # Register all .pt files in the models dir (no RAM used yet)
        count = ml_service.register_models_from_directory(Config.MODEL_DIR)
        logger.info(f"📋 {count} model(s) registered")
        # Load only the preferred default model into memory (CPU-friendly)
        default = 'sugar_bag_final'
        if default in ml_service.loaded_models:
            ml_service.switch_model(default)
        elif ml_service.loaded_models:
            # Prefer non-worldv2 models on 12GB CPU box
            preferred = [n for n in ml_service.loaded_models if 'world' not in n.lower()]
            ml_service.switch_model((preferred or list(ml_service.loaded_models))[0])
    # Start cloud sync (frames + detections → PythonAnywhere)
    try:
        from sync_client import cloud_sync
        cloud_sync.configure(lambda: camera)
        cloud_sync.start()
    except Exception as e:
        logger.warning(f'Cloud sync not started: {e}')
    print(f"""
    🚀 Edge Server Ready!
       Device: {DEVICE}
       Models: {len(ml_service.loaded_models)}
       Active: {ml_service.active_model_name or 'None'}
       Cloud:  {os.getenv('CLOUD_URL') or '(set CLOUD_URL to enable PA sync)'}
    """)

initialize()

# Start background maintenance thread (runs every 5 min, never blocks requests)
threading.Thread(target=_maintenance_loop, daemon=True, name='maintenance').start()
logger.info("🔧 Background maintenance thread started")

if __name__ == '__main__':
    port = int(os.getenv('PORT', '5000'))
    app.run(host='0.0.0.0', port=port, debug=False, threaded=True)
